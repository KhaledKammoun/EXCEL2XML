from openpyxl import load_workbook
import lxml.etree as etree


"""
XML Functions
"""
class XMLFunctions :
    # child_data : Type :  TreeNode
    @staticmethod
    def create_Item(child_data) :
        root_item = etree.Element("Item")
        element_id = etree.SubElement(root_item, "ID")
        element_id.text = str(child_data.data.id)
        element_name = etree.SubElement(root_item, "Name")
        element_name.text = child_data.data.name
        element_description = etree.SubElement(root_item, "Description")
        element_description.text = child_data.data.description
        element_children = etree.SubElement(root_item, "Children")
        return root_item
    
    @staticmethod
    def createXMLTree(root, root_xml, elements) :
        elements_xml = XMLFunctions.create_Items_List(elements)
        stack = [root, elements[0]]
        stack_xml = [root_xml, elements_xml[0]]
        root_xml.append(elements_xml[0])
        
        root.add_child(elements[0])
        for i in range(1, len(elements_xml)) :
            if elements[i - 1].data.niveau < elements[i].data.niveau :
                stack[len(stack) - 1].add_child(elements[i])
                if (len(stack_xml) == 1) :
                    stack_xml[0].append(elements_xml[i])
                else :
                    stack_xml[len(stack_xml) - 1].find("Children").append(elements_xml[i])
            else :
                while (elements[i].data.niveau <= stack[len(stack) - 1].data.niveau) :
                    stack.pop()
                    stack_xml.pop()
                stack[len(stack) - 1].add_child(elements[i])
                if (len(stack_xml) == 1) :
                    stack_xml[0].append(elements_xml[i])
                else :
                    stack_xml[len(stack_xml) - 1].find("Children").append(elements_xml[i])
            stack.append(elements[i])
            stack_xml.append(elements_xml[i])
        return root_xml
    
    @staticmethod
    def create_Items_List(elements) :
        elements_xml = []
        for element in elements :
            elements_xml.append(XMLFunctions.create_Item(element))
        return elements_xml
    
    @staticmethod
    def SaveXMLFile(root, output_path) :
        etree.ElementTree(root).write(output_path, pretty_print=True, encoding="utf-8")
        XMLFunctions.add_xml_declaration(output_path)
    
    @staticmethod
    def add_xml_declaration(xml_file):
    
        with open(xml_file, 'r', encoding='utf-8') as file:
            xml_content = file.read()

        # Add the XML declaration if it doesn't exist
        if not xml_content.startswith('<?xml'):
            xml_content = '<?xml version="1.0" encoding="UTF-8" standalone="no" ?>\n' + xml_content

            # Write the modified content back to the file
            with open(xml_file, 'w', encoding='utf-8') as file:
                file.write(xml_content)

    @staticmethod
    def createNewXMLFile(root_TreeNode, elements, output_path) :   
        root = etree.Element("BuildingInformation")
        root_xml = etree.SubElement(root, "Classification")
        system = etree.SubElement(root_xml, "System")
        name = etree.SubElement(system, "Name")
        name.text = "Archicad Classification"

        editionVersion = etree.SubElement(system, "EditionVersion")
        editionVersion.text = "v 2.0"
        editionDate = etree.SubElement(system, "EditionDate")
        year = etree.SubElement(editionDate, "Year")
        year.text = "2019"
        month = etree.SubElement(editionDate, "Month")
        month.text = "3"
        day = etree.SubElement(editionDate, "Day")
        day.text = "15"
        description = etree.SubElement(system, "Description")
        source = etree.SubElement(system, "Source")
        source.text = "www.graphisoft.com"
        items = etree.SubElement(system, "Items")

        XMLFunctions.createXMLTree(root_TreeNode, items, elements)

        XMLFunctions.SaveXMLFile(root, output_path)


def excelSheet_modulation(sheet) :

     # Convert the generator to a list for reversing
    rows_to_delete = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row))
    
    
    # Delete empty rows
    for row in reversed(rows_to_delete):
        if all(cell.value is None for cell in row):
            sheet.delete_rows(row[0].row, amount=1)

    # Convert the generator to a list for reversing
    cols_to_delete = list(sheet.iter_cols(min_col=1, max_col=sheet.max_column))
    
    # Delete empty columns
    for col in reversed(cols_to_delete):
        if all(cell.value is None for cell in col):
            sheet.delete_cols(col[0].col_idx, amount=1)
    
    return sheet


class ExcelElementsClass :
    def __init__(self, id, name, description, niveau) :
        self.id = id
        self.name = name
        self.description = description
        self.niveau = niveau
    @staticmethod
    def getAllRowsFromExcel(sheet):
        elements = []
        allRowsList = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row))
        
        # add rows value to the elements, distinct: id, name, description, niveau
        for row in allRowsList:
            elements.append(TreeNode(ExcelElementsClass(*[cell.value for cell in row][:4])))

        return elements

class TreeNode:
    def __init__(self, data):
        self.data = data
        self.children = []


    def add_child(self, child_node):
        self.children.append(child_node)

def print_tree(node, level=0, prefix="Root"):
    if level == 0:
        print(f"{prefix} - {node.data.name}")
    else:
        indent = " " * (level * 4)
        print(f"{indent}└── {node.data.niveau} - {node.data.name}")

    for child in node.children:
        print_tree(child, level + 1, f"{prefix}.{child.data.niveau}")


def createTree(root, elements) :
    # LIFO : Last In First Out ==> Stack
    stack = [root, elements[0]]
    root.add_child(elements[0])
    for i in range(1, len(elements)) :
        if elements[i - 1].data.niveau < elements[i].data.niveau :
            stack[len(stack) - 1].add_child(elements[i])
        else :
            while (elements[i].data.niveau <= stack[len(stack) - 1].data.niveau) :
                stack.pop()
            stack[len(stack) - 1].add_child(elements[i])
        stack.append(elements[i])
    return root


def createXMLFile(input_path, output_path) :
    # Create a Tree
    root = TreeNode(ExcelElementsClass('0', "Persons",None, 0))

    excelFile = load_workbook(input_path)
    workSheet = excelFile.active
    # OR :  workSheet = excelFile["Sheet1"]

    # To delete the extra empty rows and cols
    workSheet = excelSheet_modulation(workSheet)

    # contain all the excel rows with distinct value according to ('id', 'name', 'description', 'Niveau')
    elements = ExcelElementsClass.getAllRowsFromExcel(workSheet)
    createTree(root, elements)
    
    # call createNewXMLFile, using the Tree Root, as an XML element, and then save it as an xml file .
    XMLFunctions.createNewXMLFile(root, elements, output_path)
    
    
    # Print the tree starting from the root
    ## print_tree(root)

    # excelFile.save('FamilyExcel.xlsx')
    # open_file('FamilyExcel.xlsx')




from openpyxl import Workbook
import lxml.etree as etree


def return_row(item, niveau) :
    if (len(item.find("ID").text) == 8) or niveau == 3 :
        id_value = item.find("ID").text
    elif niveau == 1:
        if len(item.find("ID").text) == 2 :
            id_value = "{}|00|00".format(item.find("ID").text)
        elif len(item.find("ID").text) == 1 :
            id_value = "0{}|00|00".format(item.find("ID").text)
    elif niveau == 2 or len(item.find("ID").text) == 5:
        id_value = "{}|00".format(item.find("ID").text)
    name_value = item.find("Name").text if item.find("Name") is not None else ""

    description = item.find("Description").text

    try :
        return [id_value, name_value, description, niveau]
    except :
        return []
def convert_xml_to_excel(input_file, output_file) :
    # XML File
    tree = etree.parse(input_file)
    
    xpath_expression = "/BuildingInformation/Classification/System/Items/Item"
    items = tree.xpath(xpath_expression)

    # Create new Excel File
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Name", "Description", "Niveau"])

    for item in items :
        ws.append(return_row(item, 1))
        children = item.find("Children")
        if children is not None:
            for item_child_1 in children.findall("Item") :
                ws.append(return_row(item_child_1, 2))
                children_1 = item_child_1.find("Children")
                if children_1 is not None:
                    for item_child_2 in children_1.findall("Item") :
                        ws.append(return_row(item_child_2, 3))
                        children_2 = item_child_2.find("Children")
                        if children_2 is not None:
                            for item_child_3 in children_2.findall("Item") :
                                ws.append(return_row(item_child_3, 4))
                                children_3 = item_child_3.find("Children")
                                if children_3 is not None:
                                    for item_child_4 in children_3.findall("Item") :
                                        ws.append(return_row(item_child_4, 5))

    wb.save(output_file)
